import datetime

from openpyxl import Workbook
from openpyxl.styles import Font

def startxlsreport():
    wb = Workbook()
    ws = wb.active
    ws.title = socket.gethostname()
    st = getdt("%d-%b-%Y %H_%M_%S")
    return wb, ws, st

def getdt(fmt):
    today = datetime.today()
    return today.strftime(fmt)

def endxlsreport(wb, st):
    dt = ' from ' + st + ' to ' + getdt("%d-%b-%Y %H_%M_%S")
    fn = getbasefile() + dt + '.xlsx'
    wb.save(fn)

def headerxlsreport(ws):
    ws.cell(row=1, column=1, value="File Name")
    ws.cell(row=1, column=2, value="Full File Name")
    ws.cell(row=1, column=3, value="Folder Name")
    ws.cell(row=1, column=4, value="Date")
    ws.cell(row=1, column=5, value="Time")

    a1 = ws['A1']
    b1 = ws['B1']
    c1 = ws['C1']
    d1 = ws['D1']
    e1 = ws['E1']

    a1.font = Font(color="000000", bold=True)
    b1.font = Font(color="000000", bold=True)
    c1.font = Font(color="000000", bold=True)
    d1.font = Font(color="000000", bold=True)
    e1.font = Font(color="000000", bold=True)

def getlastrow(ws):
    rw = 1
    for cell in ws["A"]:
        if cell.value is None:
            break
        else:
            rw += 1
    return rw

def rowxlsreport(ws, fn, ffn, fld, d, t):
    rw = getlastrow(ws)

    ws.cell(row=rw, column=1, value=fn)
    ws.cell(row=rw, column=2, value=ffn)
    ws.cell(row=rw, column=3, value=fld)
    ws.cell(row=rw, column=4, value=d)
    ws.cell(row=rw, column=5, value=t)

